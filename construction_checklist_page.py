from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Dict, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font


APP_DIR = Path(__file__).resolve().parent
TEMPLATE_CANDIDATES = [
    "Copy of Draw Check List REV 12.30.25 - Jonathan.xlsx",
    "Draw Check List REV 12.30.25.xlsx",
    "Draw Check List.xlsx",
]

STATUS_OPTIONS = [
    "Pending",
    "Complete",
    "Review",
    "Missing",
    "Not Applicable",
]

AUTO_ROW_HELP = {
    2: "Loan Buyer / Capital Partner",
    3: "Next payment due",
    4: "Late payment check",
    5: "Current maturity date",
    6: "Taxes check",
    7: "Supplier code",
    8: "Property insurance",
    32: "Remaining value check",
}


def is_red_font(cell) -> bool:
    color = getattr(getattr(cell, "font", None), "color", None)
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return False
    return "FF0000" in str(rgb).upper()


def is_blue_font(cell) -> bool:
    color = getattr(getattr(cell, "font", None), "color", None)
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return False
    return "0070C0" in str(rgb).upper()


def pick_template_bytes(uploaded_file) -> Tuple[bytes | None, str | None]:
    if uploaded_file is not None:
        return uploaded_file.getvalue(), uploaded_file.name

    for candidate in TEMPLATE_CANDIDATES:
        path = APP_DIR / candidate
        if path.exists():
            return path.read_bytes(), path.name

    return None, None


@st.cache_data(show_spinner=False)
def extract_template_rows(template_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]

    section = "General"
    rows = []

    for row_idx in range(1, ws.max_row + 1):
        a = ws[f"A{row_idx}"]
        b = ws[f"B{row_idx}"]

        label = "" if a.value is None else str(a.value).strip()
        helper = "" if b.value is None else str(b.value).strip()

        if not label and not helper:
            continue

        is_section = bool(label and a.font and a.font.bold)

        if is_section:
            section = label
            continue

        rows.append(
            {
                "row_number": row_idx,
                "section": section,
                "item": label,
                "helper": helper,
                "is_red": is_red_font(a) or is_red_font(b),
                "is_blue": is_blue_font(a) or is_blue_font(b),
                "status": "Pending",
                "value": "",
                "source": "",
                "notes": "",
            }
        )

    return pd.DataFrame(rows)


def answer_from_choice(
    choice: str,
    source: str,
    good_label: str = "Complete",
    bad_label: str = "Missing",
) -> Tuple[str, str, str]:
    mapping = {
        "Current / good": (good_label, "Current", source),
        "Current": (good_label, "Current", source),
        "No issues": (good_label, "No issues", source),
        "Enough remaining value": (good_label, "Enough remaining value", source),
        "Expired / missing": (bad_label, "Expired / missing", source),
        "Delinquent": (bad_label, "Delinquent", source),
        "Late payments found": ("Review", "Late payments found", source),
        "Low / insufficient": ("Review", "Low / insufficient", source),
        "Need review": ("Review", "", source),
        "Not applicable": ("Not Applicable", "", source),
    }
    return mapping.get(choice, ("Pending", "", source))


def build_auto_answers(form_values: dict) -> Dict[int, dict]:
    answers: Dict[int, dict] = {}

    sold_status = form_values["sold_loan_status"]
    loan_buyer = form_values["loan_buyer_or_cap_partner"].strip()
    if sold_status == "Not applicable":
        answers[2] = {"status": "Not Applicable", "value": "", "source": "Salesforce later"}
    elif sold_status == "Cap partner / sold loan":
        answers[2] = {
            "status": "Complete" if loan_buyer else "Review",
            "value": loan_buyer,
            "source": "Salesforce later" if loan_buyer else "Manual review needed",
        }
    else:
        answers[2] = {"status": "Review", "value": "", "source": "Manual review needed"}

    answers[3] = {
        "status": "Complete" if form_values["next_payment_due"] else "Review",
        "value": form_values["next_payment_due"],
        "source": "Loan agreement / Salesforce later",
    }

    status, value, source = answer_from_choice(form_values["late_payment_check"], "FCI later")
    answers[4] = {"status": status, "value": value, "source": source}

    answers[5] = {
        "status": "Complete" if form_values["maturity_date"] else "Review",
        "value": form_values["maturity_date"],
        "source": "Salesforce later",
    }

    status, value, source = answer_from_choice(form_values["tax_status"], "ICE later")
    answers[6] = {"status": status, "value": value, "source": source}

    answers[7] = {
        "status": "Complete" if form_values["supplier_code"] else "Review",
        "value": form_values["supplier_code"].strip(),
        "source": "Workday later" if form_values["supplier_code"].strip() else "Manual review needed",
    }

    status, value, source = answer_from_choice(form_values["insurance_status"], "OSC later")
    answers[8] = {"status": status, "value": value, "source": source}

    status, value, source = answer_from_choice(
        form_values["remaining_value_status"],
        "Salesforce later",
        good_label="Complete",
        bad_label="Review",
    )
    answers[32] = {"status": status, "value": value, "source": source}

    return answers


def apply_auto_answers(base_df: pd.DataFrame, answers: Dict[int, dict]) -> pd.DataFrame:
    df = base_df.copy()

    for row_number, payload in answers.items():
        mask = df["row_number"] == row_number
        if not mask.any():
            continue

        for col in ["status", "value", "source"]:
            df.loc[mask, col] = payload.get(col, "")

        if row_number in AUTO_ROW_HELP:
            df.loc[mask, "notes"] = f"Starter rule: {AUTO_ROW_HELP[row_number]}"

    return df


def build_output_workbook(template_bytes: bytes, edited_rows: pd.DataFrame) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]

    ws["C1"] = "Status"
    ws["D1"] = "Value / Date"
    ws["E1"] = "Source / Notes"

    header_font = Font(bold=True, color="FF000000")
    for cell_ref in ["C1", "D1", "E1"]:
        ws[cell_ref].font = header_font

    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 36

    for _, row in edited_rows.iterrows():
        r = int(row["row_number"])
        ws[f"C{r}"] = row["status"]
        ws[f"D{r}"] = row["value"]

        note_parts = []
        if str(row.get("source", "")).strip():
            note_parts.append(str(row["source"]).strip())
        if str(row.get("notes", "")).strip():
            note_parts.append(str(row["notes"]).strip())
        ws[f"E{r}"] = " | ".join(note_parts)

        for cell_ref in [f"C{r}", f"D{r}", f"E{r}"]:
            ws[cell_ref].font = Font(color="FF000000")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def run_construction_checklist_app() -> None:
    st.title("Construction Checklist")
    st.caption(
        "Starter page for your HUD project: it reads the checklist template, focuses on the red-font items first, "
        "then lets the user review the full checklist and download an updated Excel file."
    )

    uploaded_template = st.file_uploader(
        "Upload the construction checklist template (optional if the file is already in the repo folder)",
        type=["xlsx"],
    )

    template_bytes, template_name = pick_template_bytes(uploaded_template)
    if template_bytes is None:
        st.warning("Add the checklist Excel file next to this script or upload it above.")
        st.stop()

    base_df = extract_template_rows(template_bytes)
    red_count = int(base_df["is_red"].sum())

    c1, c2, c3 = st.columns(3)
    c1.metric("Checklist items", int(len(base_df)))
    c2.metric("Red-font items", red_count)
    c3.metric("Template", template_name)

    with st.expander("Auto-fill starter inputs", expanded=True):
        left, right = st.columns(2)

        with left:
            sold_loan_status = st.selectbox(
                "NLB / sold loan status",
                ["Not applicable", "Cap partner / sold loan", "Need review"],
            )
            loan_buyer_or_cap_partner = st.text_input("Loan buyer / capital partner")
            next_payment_due = st.text_input("Next payment due")
            maturity_date = st.text_input("Current maturity date")
            supplier_code = st.text_input("Supplier code")

        with right:
            late_payment_check = st.selectbox(
                "Other payments late?",
                ["No issues", "Late payments found", "Need review"],
            )
            tax_status = st.selectbox(
                "Taxes status",
                ["Current / good", "Delinquent", "Need review"],
            )
            insurance_status = st.selectbox(
                "Property insurance",
                ["Current", "Expired / missing", "Need review"],
            )
            remaining_value_status = st.selectbox(
                "Remaining value in Salesforce",
                ["Enough remaining value", "Low / insufficient", "Need review"],
            )

    form_values = {
        "sold_loan_status": sold_loan_status,
        "loan_buyer_or_cap_partner": loan_buyer_or_cap_partner,
        "next_payment_due": next_payment_due,
        "late_payment_check": late_payment_check,
        "maturity_date": maturity_date,
        "tax_status": tax_status,
        "supplier_code": supplier_code,
        "insurance_status": insurance_status,
        "remaining_value_status": remaining_value_status,
    }

    working_df = apply_auto_answers(base_df, build_auto_answers(form_values))

    show_only_red = st.toggle("Show only red-font checklist rows first", value=True)

    editor_df = working_df.copy()
    if show_only_red:
        editor_df = editor_df[editor_df["is_red"]].copy()

    editor_df = editor_df[
        ["row_number", "section", "item", "helper", "status", "value", "source", "notes"]
    ].reset_index(drop=True)

    st.subheader("Checklist review")
    edited_df = st.data_editor(
        editor_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=["row_number", "section", "item", "helper"],
        column_config={
            "row_number": st.column_config.NumberColumn("Row", disabled=True),
            "section": st.column_config.TextColumn("Section", disabled=True),
            "item": st.column_config.TextColumn("Checklist item", width="large", disabled=True),
            "helper": st.column_config.TextColumn("Template helper text", width="large", disabled=True),
            "status": st.column_config.SelectboxColumn("Status", options=STATUS_OPTIONS, required=True),
            "value": st.column_config.TextColumn("Value / Date", width="medium"),
            "source": st.column_config.TextColumn("Source", width="medium"),
            "notes": st.column_config.TextColumn("Notes", width="large"),
        },
    )

    completed_count = int((edited_df["status"] == "Complete").sum())
    review_count = int((edited_df["status"] == "Review").sum())

    c1, c2 = st.columns(2)
    c1.metric("Completed in current view", completed_count)
    c2.metric("Needs review in current view", review_count)

    if not show_only_red:
        download_source_df = edited_df.copy()
    else:
        download_source_df = working_df.copy()
        for _, row in edited_df.iterrows():
            mask = download_source_df["row_number"] == int(row["row_number"])
            for col in ["status", "value", "source", "notes"]:
                download_source_df.loc[mask, col] = row[col]

    output_bytes = build_output_workbook(template_bytes, download_source_df)

    st.download_button(
        "Download completed checklist workbook",
        data=output_bytes,
        file_name="construction_checklist_completed.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    st.info(
        "Phase 2 is to replace the manual starter inputs with your actual Salesforce / OSC / ICE / FCI / Workday lookups."
    )


if __name__ == "__main__":
    st.set_page_config(page_title="Construction Checklist", layout="wide")
    run_construction_checklist_app()
