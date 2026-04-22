# ============================================================
# HUD Generator App — ONE FILE (Streamlit) — SF-FALLBACK READY
# Fixes requested:
# ✅ Loan ID on template = Deal Number (NOT Yardi)  -> writes Deal_Loan_Number__c into G7
# ✅ Uses your confirmed mappings AND adds robust fallbacks using the field list you provided
# ✅ Pulls from Opportunity + Property__c + Advance__c (and keeps Loan__c for servicer fallback)
# ✅ Loan Commitment prefers Advance__c.LOC_Commitment__c, then Property__c.LOC_Commitment__c, then Opp LOC_Commitment__c/Amount
# ✅ Initial Advance prefers Property__c.Initial_Disbursement_Used__c, then Property__c.Initial_Disbursement__c, then Advance__c.Initial_Disbursement_Total__c
# ✅ Total Reno Drawn prefers Property__c.Renovation_Advance_Amount_Used__c, then Advance__c.Renovation_Reserve_Total__c, then Opp Total_Amount_Advances__c (last resort)
# ✅ Interest Reserve prefers Property__c.Interest_Allocation__c, then Opp Interest_Reserves__c / Current_* fields, then Advance__c Interest reserve totals
# ✅ Borrower + Address prefer Property__c, then Opportunity/Account fallbacks
#
# NEW FIXES ADDED (for "works for me but not for them"):
# ✅ Describe cache is PER SESSION (no cross-user permission leakage)
# ✅ Permission/FLS/object access errors return empty results (doesn't crash app)
# ✅ Loan__c fetch is NON-BLOCKING (permissions won't break checks)
# ============================================================

import base64
import hashlib
import io
import json
import re
import secrets
import time
import urllib.parse
from datetime import date
from pathlib import Path
from typing import Dict, Tuple

import pandas as pd
import requests
import streamlit as st
from simple_salesforce import Salesforce
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# -----------------------------
# PAGE + STYLE
# -----------------------------
st.set_page_config(page_title="HUD Generator (Salesforce)", layout="wide")
st.markdown(
    """
<style>
  .block-container { padding-top: 1.0rem; padding-bottom: 2rem; }
  h1, h2, h3 { letter-spacing: -0.02em; }
  .soft-card {
    border: 1px solid rgba(49,51,63,.18);
    border-radius: 14px;
    padding: 14px 14px 10px 14px;
    background: rgba(255,255,255,.55);
  }
  .muted { color: rgba(49,51,63,.7); font-size: 0.92rem; }
  .big { font-size: 1.05rem; }
  .pill {
    display:inline-block; padding: 2px 10px; border-radius: 999px;
    border: 1px solid rgba(49,51,63,.18); background: rgba(255,255,255,.7);
    font-size: 0.85rem;
    margin-right: 6px;
    margin-top: 6px;
  }
  [data-testid="stTextInput"] input, [data-testid="stNumberInput"] input, [data-testid="stDateInput"] input {
    border-radius: 10px !important;
    padding: 10px 12px !important;
    font-size: 1rem !important;
  }
</style>
""",
    unsafe_allow_html=True,
)
st.title("HUD Generator App")
st.caption("Use the sidebar to switch between the HUD Generator and the Construction Checklist workflow.")
if "debug_last_sf_error" not in st.session_state:
    st.session_state.debug_last_sf_error = None

# -----------------------------
# TEMPLATE SETTINGS
# -----------------------------
APP_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = APP_DIR / "HUD TEMPLATE.xlsx"
TEMPLATE_SHEET = "TL-15255"

# Template mapping (unchanged positions; Loan ID cell gets DEAL #)
CELL_MAP = {
    "total_loan_amount": "D7",
    "initial_advance": "D8",
    "total_reno_drawn": "D9",
    "advance_amount": "D10",
    "interest_reserve": "D11",

    "deal_number": "G7",      # ✅ Loan ID cell now = Deal #
    "advance_date": "I13",

    "borrower_disp": "D14",
    "address_disp": "D15",

    "inspection_fee": "H21",
    "wire_fee": "H22",
    "construction_mgmt_fee": "H23",
    "title_fee": "H24",
}

# -----------------------------
# SECRETS
# -----------------------------
cfg = st.secrets["salesforce"]
CLIENT_ID = cfg["client_id"]
AUTH_HOST = cfg.get("auth_host", "https://cvest.my.salesforce.com").rstrip("/")
REDIRECT_URI = cfg["redirect_uri"].rstrip("/")
CLIENT_SECRET = cfg.get("client_secret")
AUTH_URL = f"{AUTH_HOST}/services/oauth2/authorize"
TOKEN_URL = f"{AUTH_HOST}/services/oauth2/token"

# -----------------------------
# PKCE HELPERS
# -----------------------------
def b64url_no_pad(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).rstrip(b"=").decode("utf-8")

def make_verifier() -> str:
    v = secrets.token_urlsafe(96)
    return v[:128]

def make_challenge(verifier: str) -> str:
    return b64url_no_pad(hashlib.sha256(verifier.encode("utf-8")).digest())

@st.cache_resource
def pkce_store():
    return {}

store = pkce_store()
now = time.time()
TTL = 900
for s, (_v, t0) in list(store.items()):
    if now - t0 > TTL:
        store.pop(s, None)

# -----------------------------
# UTIL
# -----------------------------
def soql_quote(s: str) -> str:
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'") + "'"

def digits_only(x: str) -> str:
    return re.sub(r"\D", "", x or "")

def normalize_text(x):
    return "" if x is None else str(x).strip()

def pick_first(*vals):
    for v in vals:
        if v is None:
            continue
        s = str(v).strip()
        if s != "":
            return s
    return ""

def parse_money(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if s == "":
        return 0.0
    s = s.replace("$", "").replace(",", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    try:
        x = float(s)
        return -x if neg else x
    except Exception:
        return 0.0

def fmt_money(x) -> str:
    try:
        return f"${float(x):,.2f}"
    except Exception:
        return "$0.00"

def parse_date_any(x):
    if x in ("", None):
        return None
    dt = pd.to_datetime(x, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()

def fmt_date_mmddyyyy(x) -> str:
    d = parse_date_any(x)
    return d.strftime("%m/%d/%Y") if d else ""

def norm(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", "_", regex=True)
        .str.replace(r"[^0-9a-z_]+", "", regex=True)
    )
    return df

def extract_order_id_deal_prefix(order_id_val: str) -> str:
    if not order_id_val:
        return ""
    left = str(order_id_val).split("-", 1)[0].strip()
    return digits_only(left)

def strip_zip4(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"(\b\d{5})-\d{4}\b", r"\1", str(s))

DIR_MAP = {
    "north": "n", "n": "n",
    "south": "s", "s": "s",
    "east": "e", "e": "e",
    "west": "w", "w": "w",
    "northeast": "ne", "ne": "ne",
    "northwest": "nw", "nw": "nw",
    "southeast": "se", "se": "se",
    "southwest": "sw", "sw": "sw",
}
STATE_MAP = {"oregon": "or", "or": "or", "washington": "wa", "wa": "wa", "california": "ca", "ca": "ca"}
SUFFIX_MAP = {
    "street": "st", "st": "st",
    "avenue": "ave", "ave": "ave",
    "road": "rd", "rd": "rd",
    "drive": "dr", "dr": "dr",
    "lane": "ln", "ln": "ln",
    "court": "ct", "ct": "ct",
    "place": "pl", "pl": "pl",
    "terrace": "ter", "ter": "ter",
    "trail": "trl", "trl": "trl",
    "circle": "cir", "cir": "cir",
    "boulevard": "blvd", "blvd": "blvd",
    "parkway": "pkwy", "pkwy": "pkwy",
}

def address_tokens(s: str) -> set:
    if not s:
        return set()
    s = strip_zip4(str(s)).lower()
    s = re.sub(r"[,#]", " ", s)
    s = s.replace("-", " ")
    s = re.sub(r"[^0-9a-z\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    toks = s.split()
    out = []
    for t in toks:
        if t in DIR_MAP:
            out.append(DIR_MAP[t])
        elif t in STATE_MAP:
            out.append(STATE_MAP[t])
        elif t in SUFFIX_MAP:
            out.append(SUFFIX_MAP[t])
        else:
            out.append(t)
    return set(out)

def zip5_from_addr(s: str) -> str:
    s = strip_zip4(s or "")
    m = re.search(r"\b(\d{5})\b", s)
    return m.group(1) if m else ""

def house_num_from_addr(s: str) -> str:
    m = re.match(r"\s*(\d+)\b", (s or "").strip())
    return m.group(1) if m else ""

def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0

def pick_first_nonblank_field(record: dict, fields: list):
    """
    Returns (field_name, value) for first field with nonblank value.
    """
    if not record:
        return None, None
    for f in fields:
        if f in record:
            v = record.get(f)
            if v is None:
                continue
            s = str(v).strip()
            if s != "":
                return f, v
    return None, None

# -----------------------------
# OAUTH FLOW (PKCE)
# -----------------------------
qp = st.query_params
code = qp.get("code")
state = qp.get("state")
err = qp.get("error")
err_desc = qp.get("error_description")

if err:
    st.error(f"Login error: {err}")
    if err_desc:
        st.code(err_desc)
    st.stop()

if "sf_token" not in st.session_state:
    st.session_state.sf_token = None

def exchange_code_for_token(code: str, verifier: str) -> dict:
    data = {
        "grant_type": "authorization_code",
        "client_id": CLIENT_ID,
        "redirect_uri": REDIRECT_URI,
        "code": code,
        "code_verifier": verifier,
    }
    if CLIENT_SECRET:
        data["client_secret"] = CLIENT_SECRET
    resp = requests.post(TOKEN_URL, data=data, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f"Token exchange failed ({resp.status_code}): {resp.text}")
    return resp.json()

if code:
    if not state or state not in store:
        st.error("Login link expired. Click login again.")
        st.stop()
    verifier, _t0 = store.pop(state)
    tok = exchange_code_for_token(code, verifier)
    st.session_state.sf_token = tok
    st.query_params.clear()
    st.rerun()

if not st.session_state.sf_token:
    new_state = secrets.token_urlsafe(24)
    new_verifier = make_verifier()
    new_challenge = make_challenge(new_verifier)
    store[new_state] = (new_verifier, time.time())

    login_params = {
        "response_type": "code",
        "client_id": CLIENT_ID,
        "redirect_uri": REDIRECT_URI,
        "code_challenge": new_challenge,
        "code_challenge_method": "S256",
        "state": new_state,
        "prompt": "login",
        "scope": "api refresh_token",
    }
    login_url = AUTH_URL + "?" + urllib.parse.urlencode(login_params)
    st.info("Step 1: Log in.")
    st.link_button("Login", login_url)
    st.stop()

tok = st.session_state.sf_token
access_token = tok.get("access_token")
instance_url = tok.get("instance_url")

if not access_token or not instance_url:
    st.error("Login token missing needed values.")
    st.stop()

sf = Salesforce(instance_url=instance_url, session_id=access_token)

topc1, topc2 = st.columns([3, 1])
with topc1:
    st.success("✅ Logged in")
    st.caption(f"Connected to: {instance_url}")
with topc2:
    if st.button("Log out"):
        st.session_state.sf_token = None
        st.rerun()

# -----------------------------
# LOAD EXCEL CHECK FILES
# -----------------------------
OSC_CANDIDATES = [
    "OSC_Zstatus_COREVEST_2026-04-14_202850.xlsx",
    "OSC_Zstatus_COREVEST_2026-03-24_064223.xlsx",
]
CAF_CANDIDATES = [
    "TAXES_CAF National 52874_3.9.26.xlsx",
    "Corevest_CAF National 52874_3.9.26.xlsx",
]

def first_existing_path(candidates):
    for c in candidates:
        p1 = Path(c)
        if p1.exists():
            return str(p1)
        p2 = Path("/mnt/data") / c
        if p2.exists():
            return str(p2)
    return candidates[0]

@st.cache_data(show_spinner=False)
def load_osc_excel():
    path = first_existing_path(OSC_CANDIDATES)
    try:
        x = pd.read_excel(path, sheet_name=None, dtype=str)
        df = x["COREVEST"] if "COREVEST" in x else list(x.values())[0]
        df = norm(df)
        return df, path, None
    except Exception as e:
        return pd.DataFrame(), path, str(e)

@st.cache_data(show_spinner=False)
def load_caf_excel():
    path = first_existing_path(CAF_CANDIDATES)
    try:
        x = pd.read_excel(path, sheet_name=None, dtype=str)
        df = x["Completed"] if "Completed" in x else list(x.values())[0]
        df = norm(df)
        return df, path, None
    except Exception as e:
        return pd.DataFrame(), path, str(e)

osc_df, osc_path_used, osc_err = load_osc_excel()
caf_df, caf_path_used, caf_err = load_caf_excel()

# -----------------------------
# DESCRIBE CACHES (FIXED: PER SESSION)
# -----------------------------
# IMPORTANT: st.cache_resource here would be shared across users.
# We want describe results per user session because permissions differ.
if "DESC" not in st.session_state:
    st.session_state.DESC = {}
DESC = st.session_state.DESC

def get_obj_fields(obj_name: str) -> set:
    if obj_name in DESC:
        return DESC[obj_name]
    try:
        d = sf.__getattr__(obj_name).describe()
        fields = {f.get("name") for f in d.get("fields", []) if f.get("name")}
        DESC[obj_name] = fields
        return fields
    except Exception:
        DESC[obj_name] = set()
        return set()

def filter_existing_fields(obj_name: str, fields: list) -> list:
    existing = get_obj_fields(obj_name)
    if not existing:
        return fields
    return [f for f in fields if f in existing]

def choose_first_existing(obj_name: str, candidates: list):
    existing = get_obj_fields(obj_name)
    if not existing:
        return None
    for c in candidates:
        if c in existing:
            return c
    return None

# -----------------------------
# SAFE QUERY (FIXED: PERMISSION ERRORS DON'T CRASH)
# -----------------------------
def sf_query_all(sf: Salesforce, soql: str):
    return sf.query_all(soql).get("records", [])

def _is_perm_error(msg: str) -> bool:
    m = (msg or "").lower()
    needles = [
        "insufficient", "permission", "not authorized", "not permitted",
        "invalid_type",  # no access to object (or object doesn't exist for user)
        "insufficient_access", "insufficient access",
        "insufficient_privileges",
        "insufficient_access_on_cross_reference_entity",
        "entity is not accessible",
        "field is not accessible",
        "no access", "access denied",
    ]
    return any(n in m for n in needles)

def try_query_drop_missing(sf: Salesforce, obj_name: str, fields, where_clause: str, limit=200, order_by=None):
    fields = list(dict.fromkeys([f for f in fields if f]))
    fields = filter_existing_fields(obj_name, fields)

    # If describe failed or they have no accessible fields, don't crash app
    if not fields:
        st.session_state.debug_last_sf_error = {
            "soql": f"(no accessible fields) FROM {obj_name}",
            "error": "No accessible fields for this user (object/FLS).",
        }
        return [], [], ""

    if order_by:
        ob_field = order_by.split()[0].strip()
        existing = get_obj_fields(obj_name)
        if existing and ob_field not in existing:
            order_by = None

    while True:
        soql = f"SELECT {', '.join(fields)} FROM {obj_name} WHERE {where_clause}"
        if order_by:
            soql += f" ORDER BY {order_by}"
        soql += f" LIMIT {int(limit)}"
        try:
            rows = sf_query_all(sf, soql)
            return rows, fields, soql
        except Exception as e:
            msg = str(e)
            st.session_state.debug_last_sf_error = {"soql": soql, "error": msg}

            # FIX: Treat permission/FLS/sharing errors as "no rows" instead of crashing
            if _is_perm_error(msg):
                return [], fields, soql

            if order_by and ("unexpected token" in msg.lower() or "nulls" in msg.lower()):
                order_by = None
                continue

            m1 = re.search(r"No such column '([^']+)'", msg)
            m3 = re.search(r"Invalid field: ([^,]+)", msg)
            m4 = re.search(r"INVALID_FIELD: ([^:]+):", msg)
            bad = None
            if m1:
                bad = m1.group(1).strip()
            elif m3:
                bad = m3.group(1).strip()
            elif m4:
                bad = m4.group(1).strip()

            if bad and bad in fields:
                fields.remove(bad)
                if not fields:
                    # FIX: don't crash whole app
                    return [], [], soql
                continue

            raise RuntimeError("Salesforce query failed.") from e

# -----------------------------
# SF FETCHES
# -----------------------------
def fetch_opportunity_by_deal_number(deal_number: str):
    dn_digits = digits_only((deal_number or "").strip())
    if not dn_digits:
        return None

    opp_fields = [
        "Id", "Name",
        "Deal_Loan_Number__c",
        "Account_Name__c",
        "CloseDate",
        "Servicer_Commitment_Id__c",
        "Servicer_Status__c",
        "Next_Payment_Date__c",
        "Late_Fees_Servicer__c",

        # Fallback monetary fields (from your list)
        "Amount",
        "LOC_Commitment__c",
        "Current_Loan_Amount__c",
        "Final_Loan_Amount__c",
        "Total_Amount_Advances__c",

        # Interest reserve fallbacks
        "Current_Interest_Reserves_Paid__c",
        "Current_Interest_Reserves_Remaining__c",
        "Interest_Reserves__c",
        "Current_UPB_Interest_Reserves__c",
        "Current_UPB_Interest_Reserve__c",
    ]

    where = (
        "("
        f"Deal_Loan_Number__c = {soql_quote(dn_digits)}"
        f" OR Deal_Loan_Number__c LIKE {soql_quote('%' + dn_digits + '%')}"
        ")"
    )
    rows, _used, _soql = try_query_drop_missing(sf, "Opportunity", opp_fields, where, limit=10, order_by="CloseDate DESC")
    if not rows:
        return None
    r = rows[0].copy()
    r.pop("attributes", None)
    return r

def fetch_property_for_deal(opp_id: str):
    lk = choose_first_existing("Property__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId"])
    if not lk:
        return None

    prop_fields = [
        "Id", "Name", lk,
        "Servicer_Id__c",
        "Full_Address__c",
        "Borrower_Name__c",
        "Yardi_Id__c",

        # From your list
        "Initial_Disbursement_Used__c",
        "Initial_Disbursement__c",
        "Initial_Disbursement_Total__c",
        "Initial_Disbursement_Remaining__c",
        "Total_Initial_Disbursement__c",

        "Interest_Allocation__c",

        # Total loan amount fallbacks
        "LOC_Commitment__c",
        "Outstanding_Facility_Amount__c",
        "Current_Outstanding_Loan_Amount__c",
        "Max_Total_Loan_Amount__c",
        "Max_Total_Loan_Amount_Input__c",

        # Total Reno drawn / reserve funded fallbacks (from your list)
        "Renovation_Advance_Amount_Used__c",
        "Approved_Renovation_Holdback__c",
        "Renovation_Reserve_Total__c",  # (may not exist on Property__c; drop-missing loop handles)
        "Interest_Reserves__c",         # listed under "Total Reno Drawn" group in your export

        # Requested/funding date fallbacks
        "Requested_Funding_Date__c",
        "Funding_Date__c",
        "First_Funding_Date__c",

        # Other useful values
        "Holdback_To_Rehab_Ratio__c",
        "Late_Fees_Servicer__c",
    ]

    where = f"{lk} = {soql_quote(opp_id)}"
    try:
        rows, _used, _soql = try_query_drop_missing(sf, "Property__c", prop_fields, where, limit=5, order_by="CreatedDate DESC")
        if not rows:
            return None
        r = rows[0].copy()
        r.pop("attributes", None)
        return r
    except Exception:
        st.warning("⚠️ Could not pull property details. Continuing without them.")
        return None

def fetch_loan_for_deal(opp_id: str):
    """
    FIX: Non-blocking. If user doesn't have Loan__c access/FLS, this returns None.
    """
    lk = choose_first_existing("Loan__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId"])
    if not lk:
        return None

    loan_fields = ["Id", "Name", lk, "Servicer_Loan_Status__c", "Servicer_Loan_Id__c", "Next_Payment_Date__c"]
    where = f"{lk} = {soql_quote(opp_id)}"
    try:
        rows, _used, _soql = try_query_drop_missing(sf, "Loan__c", loan_fields, where, limit=5, order_by="CreatedDate DESC")
    except Exception:
        # extra safety; should be rare now
        return None

    if not rows:
        return None
    r = rows[0].copy()
    r.pop("attributes", None)
    return r

def fetch_advances_for_deal(opp_id: str):
    """
    Pull multiple advances; we will choose values using best "nonblank" priority.
    """
    lk = choose_first_existing("Advance__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId", "Advance__c"])
    if not lk:
        return []

    adv_fields = [
        "Id", "Name", lk,

        # Amounts and key fields from your list
        "LOC_Commitment__c",
        "Advance__c",  # (reference)
        "Approved_Advance_Amount_Total__c",
        "Approved_Advance_Amount_Max_Total__c",
        "Renovation_Reserve_Total__c",
        "Initial_Disbursement_Total__c",

        # Interest reserve totals
        "Interest_Reserve_Total__c",
        "Interest_Reserve_Subtotal__c",
        "Total_Interest_Reserves_andStub_Interest__c",
        "Remaining_Interest_Reserve__c",

        # Dates
        "Target_Advance_Date__c",
        "Wire_Date__c",
        "Date_Advance_Requested__c",
    ]

    where = f"{lk} = {soql_quote(opp_id)}"
    try:
        rows, _used, _soql = try_query_drop_missing(sf, "Advance__c", adv_fields, where, limit=50, order_by="CreatedDate DESC")
        cleaned = []
        for r in rows:
            rr = r.copy()
            rr.pop("attributes", None)
            cleaned.append(rr)
        return cleaned
    except Exception:
        return []

# -----------------------------
# OFFLINE LOOKUPS (OSC + CAF)
# -----------------------------
def osc_lookup(servicer_key: str):
    if osc_df.empty:
        return {"found": False, "error": "Insurance file did not load.", "row": None}
    if "account_number" not in osc_df.columns:
        return {"found": False, "error": "Insurance file is missing the ID field.", "row": None}
    key = (servicer_key or "").strip()
    if not key:
        return {"found": False, "error": "Missing servicer ID.", "row": None}
    hit = osc_df[osc_df["account_number"].astype(str).str.strip() == key]
    if hit.empty:
        return {"found": False, "error": "No insurance record found for that servicer ID.", "row": None}
    return {"found": True, "error": None, "row": hit.iloc[0].to_dict()}

def caf_try_match_by_deal_id(deal_digits: str):
    if caf_df.empty:
        return {"found": False, "error": "Payment file did not load.", "row": None, "method": "deal id"}
    if "order_id" not in caf_df.columns:
        return {"found": False, "error": "Payment file is missing deal IDs.", "row": None, "method": "deal id"}
    dn = digits_only(deal_digits)
    if not dn:
        return {"found": False, "error": "Missing deal number.", "row": None, "method": "deal id"}
    prefixes = caf_df["order_id"].astype(str).fillna("").map(extract_order_id_deal_prefix)
    hit = caf_df[prefixes == dn]
    if hit.empty:
        return {"found": False, "error": "No payment record found by deal ID.", "row": None, "method": "deal id"}
    return {"found": True, "error": None, "row": hit.iloc[0].to_dict(), "method": "deal id"}

def caf_try_match_by_address(sf_addr: str, osc_addr: str):
    if caf_df.empty:
        return {"found": False, "error": "Payment file did not load.", "row": None, "method": "address"}
    if "property_address" not in caf_df.columns:
        return {"found": False, "error": "Payment file is missing property addresses.", "row": None, "method": "address"}
    target = normalize_text(sf_addr) or normalize_text(osc_addr)
    if not target:
        return {"found": False, "error": "No address available to match.", "row": None, "method": "address"}

    target_zip = zip5_from_addr(target)
    target_house = house_num_from_addr(target)
    target_tokens = address_tokens(target)

    caf_addr_raw = caf_df["property_address"].astype(str).fillna("")
    candidates = caf_df.copy()
    candidates["_addr_raw"] = caf_addr_raw

    if target_zip:
        candidates["_zip5"] = candidates["_addr_raw"].map(zip5_from_addr)
        candidates = candidates[candidates["_zip5"] == target_zip]
    if target_house and not candidates.empty:
        candidates["_house"] = candidates["_addr_raw"].map(house_num_from_addr)
        candidates = candidates[candidates["_house"] == target_house]

    if candidates.empty:
        candidates = caf_df.copy()
        candidates["_addr_raw"] = caf_addr_raw

    scores = []
    for idx, row in candidates.iterrows():
        toks = address_tokens(row["_addr_raw"])
        score = jaccard(target_tokens, toks)
        scores.append((idx, score))
    if not scores:
        return {"found": False, "error": "No address candidates found.", "row": None, "method": "address"}

    best_idx, best_score = max(scores, key=lambda t: t[1])
    if best_score < 0.45:
        return {"found": False, "error": "No close address match found.", "row": None, "method": "address"}
    return {"found": True, "error": None, "row": caf_df.loc[best_idx].to_dict(), "method": "address match"}

def pick_payment_statuses(caf_row: dict):
    out = []
    if not caf_row:
        return out
    for col in ["inst_1_payment_status", "inst_2_payment_status", "inst_3_payment_status", "inst_4_payment_status"]:
        if col in caf_row:
            v = normalize_text(caf_row.get(col))
            if v:
                out.append((col, v))
    if not out:
        for k, v in caf_row.items():
            if "payment_status" in (k or "") and v:
                out.append((k, normalize_text(v)))
    return out

def is_payment_status_ok(val: str) -> bool:
    t = (val or "").strip().lower()
    if t == "":
        return False
    bad_words = ["delinquent", "late", "unpaid", "past due", "default", "foreclosure"]
    return not any(w in t for w in bad_words)

# -----------------------------
# PRECHECKS (OLD LOGIC STYLE)
# -----------------------------
TARGET_INSURANCE_OK = "outside policy in-force"

def run_prechecks(opp: dict, prop: dict, loan: dict, user_deal_input: str):
    deal_digits = digits_only(user_deal_input)
    deal_label = normalize_text(opp.get("Deal_Loan_Number__c")) or user_deal_input

    deal_name = normalize_text(opp.get("Name"))
    acct_name = normalize_text(opp.get("Account_Name__c"))

    servicer_key = pick_first(
        (prop or {}).get("Servicer_Id__c"),
        (opp or {}).get("Servicer_Commitment_Id__c"),
        (loan or {}).get("Servicer_Loan_Id__c"),
    )

    # System address (display only)
    sf_full_address = normalize_text((prop or {}).get("Full_Address__c"))
    sf_full_address_disp = sf_full_address.upper() if sf_full_address else ""

    # OSC (required / blocking)
    osc = osc_lookup(servicer_key)
    osc_primary = ""
    osc_ok = False
    osc_addr_disp = ""
    if osc["found"]:
        r = osc["row"] or {}
        osc_primary = normalize_text(r.get("primary_status"))
        osc_ok = (osc_primary.strip().lower() == TARGET_INSURANCE_OK)
        osc_addr_disp = " ".join([
            normalize_text(r.get("property_street")),
            normalize_text(r.get("property_city")),
            normalize_text(r.get("property_state")),
            normalize_text(r.get("property_zip")),
        ]).strip().upper()

    # CAF (optional / best-effort)
    caf = caf_try_match_by_deal_id(deal_digits)
    if not caf.get("found"):
        caf = caf_try_match_by_address(sf_full_address, osc_addr_disp)

    caf_addr_disp = ""
    caf_statuses = []
    caf_ok = False
    caf_found = bool(caf.get("found"))
    if caf_found:
        row = caf.get("row") or {}
        caf_addr_disp = normalize_text(row.get("property_address")).upper()
        caf_statuses = pick_payment_statuses(row)
        if caf_statuses:
            caf_ok = all(is_payment_status_ok(v) for (_k, v) in caf_statuses)

    # ONLY OSC blocks (old behavior)
    osc_blocking_ok = bool(servicer_key) and osc.get("found") and osc_ok
    overall_ok = bool(osc_blocking_ok)

    checks = []
    if not servicer_key:
        checks.append({"Check":"Servicer identifier","Value":"(missing)","Result":"Stop","Note":"Missing identifier needed to find the insurance record."})
    elif not osc.get("found"):
        checks.append({"Check":"Insurance status","Value":osc.get("error","Not found"),"Result":"Stop","Note":"We need an insurance record before creating the HUD."})
    else:
        checks.append({"Check":"Insurance status","Value":osc_primary if osc_primary else "(blank)","Result":"OK" if osc_ok else "Stop","Note":"Must be outside-policy in-force."})

    if caf_found:
        checks.append({"Check":"Payment info (optional)","Value":"Found","Result":"OK" if caf_ok else "Review","Note":"Shown for visibility; it does not block HUD creation."})
    else:
        checks.append({"Check":"Payment info (optional)","Value":caf.get("error","Not found"),"Result":"Review","Note":"Not required to create the HUD."})

    # HUD address
    hud_address_disp = osc_addr_disp or sf_full_address_disp
    checks.append({"Check":"HUD address source","Value":"Insurance record" if osc_addr_disp else "System address","Result":"OK" if hud_address_disp else "Review","Note":"HUD uses insurance address when available."})

    return {
        "deal_number": deal_label,
        "deal_name": deal_name,
        "account_name": acct_name,
        "servicer_key": servicer_key,
        "checks": checks,
        "overall_ok": overall_ok,
        "sf_address": sf_full_address_disp,
        "osc_address": osc_addr_disp,
        "caf_address": caf_addr_disp,
        "caf_method": caf.get("method", ""),
        "hud_address_disp": hud_address_disp,
    }

# -----------------------------
# EXCEL TEMPLATE OUTPUT
# -----------------------------
def _is_red_font(cell) -> bool:
    c = getattr(getattr(cell, "font", None), "color", None)
    rgb = getattr(c, "rgb", None)
    if not rgb:
        return False
    rgb = str(rgb).upper()
    return "FF0000" in rgb

def _clear_red_text(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, "") and _is_red_font(cell):
                cell.value = None

def build_hud_excel_bytes_from_template(ctx: dict) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError("HUD template not found. Add it to your repo next to app.py.")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active
    _clear_red_text(ws)

    def write_cell(key, value):
        addr = CELL_MAP.get(key)
        if not addr:
            return
        ws[addr] = value

    # TEXT
    write_cell("deal_number", str(ctx.get("deal_number", "")))      # ✅ Loan ID = Deal #
    write_cell("advance_date", str(ctx.get("advance_date", "")))
    write_cell("borrower_disp", str(ctx.get("borrower_disp", "")))
    write_cell("address_disp", str(ctx.get("address_disp", "")))

    # NUMBERS
    write_cell("total_loan_amount", float(ctx.get("total_loan_amount", 0.0)))
    write_cell("initial_advance", float(ctx.get("initial_advance", 0.0)))
    write_cell("total_reno_drawn", float(ctx.get("total_reno_drawn", 0.0)))
    write_cell("advance_amount", float(ctx.get("advance_amount", 0.0)))
    write_cell("interest_reserve", float(ctx.get("interest_reserve", 0.0)))

    write_cell("inspection_fee", float(ctx.get("inspection_fee", 0.0)))
    write_cell("wire_fee", float(ctx.get("wire_fee", 0.0)))
    write_cell("construction_mgmt_fee", float(ctx.get("construction_mgmt_fee", 0.0)))
    write_cell("title_fee", float(ctx.get("title_fee", 0.0)))

    black = Font(color="FF000000")
    for addr in set(CELL_MAP.values()):
        try:
            ws[addr].font = ws[addr].font.copy(color=black.color)
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# -----------------------------
# SESSION DEFAULTS
# -----------------------------

# -----------------------------

def ensure_default(key, val):
    if key not in st.session_state:
        st.session_state[key] = val

def run_hud_generator_page():
    ensure_default("deal_number_input", "")
    ensure_default("precheck_ran", False)
    ensure_default("precheck_payload", None)
    ensure_default("allow_override", False)

    ensure_default("inp_advance_amount", "")
    ensure_default("inp_holdback_pct", "")
    ensure_default("inp_advance_date", date.today())
    ensure_default("inp_inspection_fee", "")
    ensure_default("inp_wire_fee", "")
    ensure_default("inp_construction_mgmt_fee", "")
    ensure_default("inp_title_fee", "")

    # -----------------------------
    # Troubleshooting expander
    # -----------------------------
    with st.expander("Troubleshooting (optional)", expanded=False):
        st.write("Insurance file:", osc_path_used, "✅" if osc_err is None else "❌")
        if osc_err:
            st.code(osc_err)
        st.write("Payment file:", caf_path_used, "✅" if caf_err is None else "❌")
        if caf_err:
            st.code(caf_err)
        st.write("HUD template loaded:", "✅" if TEMPLATE_PATH.exists() else "❌")
        if st.session_state.debug_last_sf_error:
            st.markdown("Salesforce error details:")
            st.code(st.session_state.debug_last_sf_error.get("soql", ""))
            st.code(st.session_state.debug_last_sf_error.get("error", ""))

    # -----------------------------
    # UI — DEAL INPUT + PRECHECKS
    # -----------------------------
    st.markdown('<div class="soft-card">', unsafe_allow_html=True)
    c1, c2 = st.columns([2.4, 1.2])
    with c1:
        deal_input = st.text_input("Deal Number", key="deal_number_input", placeholder="Type the deal number")
    with c2:
        run_btn = st.button("Run checks", type="primary", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if run_btn:
        st.session_state.precheck_ran = False
        st.session_state.precheck_payload = None
        st.session_state.allow_override = False

        with st.spinner("Finding your deal..."):
            opp = fetch_opportunity_by_deal_number(deal_input)

        if not opp:
            st.error("No deal found for that number. Double-check the deal number and try again.")
            st.stop()

        opp_id = opp.get("Id")

        # FIX: Loan__c is non-blocking (permissions won't crash whole app)
        with st.spinner("Pulling related info..."):
            prop = fetch_property_for_deal(opp_id) if opp_id else None
            try:
                loan = fetch_loan_for_deal(opp_id) if opp_id else None
            except Exception:
                loan = None
                st.warning("⚠️ Could not pull Loan details (permissions). Continuing without Loan__c.")
            advances = fetch_advances_for_deal(opp_id) if opp_id else []

        with st.spinner("Running checks..."):
            payload = run_prechecks(opp, prop, loan, deal_input)

        st.session_state.precheck_payload = {"opp": opp, "prop": prop, "loan": loan, "advances": advances, "payload": payload}
        st.session_state.precheck_ran = True

    # -----------------------------
    # SHOW CHECK RESULTS + ADDRESS VIEW
    # -----------------------------
    if st.session_state.precheck_ran and st.session_state.precheck_payload:
        opp = st.session_state.precheck_payload["opp"]
        prop = st.session_state.precheck_payload.get("prop") or {}
        payload = st.session_state.precheck_payload["payload"]

        st.subheader("Check results")
        st.markdown(
            f"""
    <div class="soft-card">
      <div class="big"><b>{payload['deal_number']}</b> — {payload['deal_name']}</div>
      <div class="muted">{payload['account_name']}</div>
      <div style="margin-top:8px;">
        <span class="pill">Servicer Identifier: <b>{payload['servicer_key'] if payload['servicer_key'] else '—'}</b></span>
        <span class="pill">Borrower (SF): <b>{(prop.get('Borrower_Name__c') or '') or '—'}</b></span>
      </div>
    </div>
    """,
            unsafe_allow_html=True,
        )

        df_checks = pd.DataFrame(payload["checks"])[["Check", "Value", "Result", "Note"]]
        st.dataframe(df_checks, use_container_width=True, hide_index=True)

        st.markdown("### Address comparison")
        a1, a2, a3 = st.columns(3)
        with a1:
            st.markdown("**Address from our system**")
            st.code(payload.get("sf_address") or "(blank)")
        with a2:
            st.markdown("**Insurance record address (used on HUD)**")
            st.code(payload.get("osc_address") or "(blank)")
        with a3:
            st.markdown("**Payment record address (optional)**")
            st.code(payload.get("caf_address") or "(blank)")
            st.caption(f"How it matched: {payload.get('caf_method') or '(not matched)'}")

        if payload["overall_ok"]:
            st.success("✅ Required checks passed. You can continue to build the HUD.")
            st.session_state.allow_override = True
        else:
            st.error("🚫 Required checks did not pass — HUD should NOT be created yet.")
            st.session_state.allow_override = st.checkbox("Override and continue anyway", value=False)

    # -----------------------------
    # HUD INPUTS (ONLY AFTER REQUIRED CHECKS)
    # -----------------------------
    if st.session_state.precheck_ran and st.session_state.precheck_payload and st.session_state.allow_override:
        opp = st.session_state.precheck_payload["opp"]
        prop = st.session_state.precheck_payload.get("prop") or {}
        advances = st.session_state.precheck_payload.get("advances") or []
        payload = st.session_state.precheck_payload["payload"]

        borrower_default = (pick_first(prop.get("Borrower_Name__c"), opp.get("Account_Name__c")) or "").strip().upper()
        address_default = payload.get("hud_address_disp") or ""

        st.subheader("HUD inputs")
        st.caption("Type amounts like `1200` or `$1,200` (leave blank for $0). Dates are mm/dd/yyyy.")

        with st.form("hud_form", clear_on_submit=False):
            cA, cB, cC = st.columns([1.2, 1.0, 1.2])

            with cA:
                st.markdown("**Borrower info**")
                borrower_val = st.text_input("Borrower (for the form)", value=borrower_default, key="inp_borrower_disp")
                addr_val = st.text_input("Address (for the form)", value=address_default, key="inp_address_disp")

            with cB:
                st.markdown("**Advance**")
                adv_amt_raw = st.text_input("Advance Amount", key="inp_advance_amount", placeholder="e.g., 25000")
                holdback_pct = st.text_input("Holdback % (optional)", key="inp_holdback_pct", placeholder="e.g., 20%")
                adv_date = st.date_input("Advance Date", key="inp_advance_date")

            with cC:
                st.markdown("**Fees**")
                insp_raw = st.text_input("3rd party Inspection Fee", key="inp_inspection_fee", placeholder="leave blank for 0")
                wire_raw = st.text_input("Wire Fee", key="inp_wire_fee", placeholder="leave blank for 0")
                cm_raw = st.text_input("Construction Management Fee", key="inp_construction_mgmt_fee", placeholder="leave blank for 0")
                title_raw = st.text_input("Title Fee", key="inp_title_fee", placeholder="leave blank for 0")

            submitted = st.form_submit_button("Build HUD Excel", type="primary", use_container_width=True)

        if submitted:
            advance_amount = parse_money(adv_amt_raw)
            inspection_fee = parse_money(insp_raw)
            wire_fee = parse_money(wire_raw)
            construction_mgmt_fee = parse_money(cm_raw)
            title_fee = parse_money(title_raw)

            hb = (holdback_pct or "").strip()
            if hb and not hb.endswith("%"):
                try:
                    v = float(hb.replace("%", "").strip())
                    hb = f"{v:.0f}%"
                except Exception:
                    pass

            # -----------------------------
            # FALLBACK LOGIC USING YOUR FIELD LIST
            # -----------------------------
            # Total Loan Amount (Commitment): Advance__c.LOC_Commitment__c -> Property__c.LOC_Commitment__c -> Opp LOC_Commitment__c -> Opp Amount
            adv_loc_val = None
            for a in advances:
                _f, adv_loc_val = pick_first_nonblank_field(a, ["LOC_Commitment__c"])
                if adv_loc_val is not None:
                    break
            total_loan_amount_val = pick_first(
                adv_loc_val,
                prop.get("LOC_Commitment__c"),
                opp.get("LOC_Commitment__c"),
                opp.get("Final_Loan_Amount__c"),
                opp.get("Current_Loan_Amount__c"),
                opp.get("Amount"),
            )
            sf_total_loan_amount = parse_money(total_loan_amount_val)

            # Initial Advance: Property__c.Initial_Disbursement_Used__c -> Property__c.Initial_Disbursement__c -> Advance__c.Initial_Disbursement_Total__c
            adv_init_val = None
            for a in advances:
                _f, adv_init_val = pick_first_nonblank_field(a, ["Initial_Disbursement_Total__c"])
                if adv_init_val is not None:
                    break
            initial_advance_val = pick_first(
                prop.get("Initial_Disbursement_Used__c"),
                prop.get("Initial_Disbursement__c"),
                prop.get("Total_Initial_Disbursement__c"),
                adv_init_val,
            )
            sf_initial_advance = parse_money(initial_advance_val)

            # Total Reno Drawn: Property__c.Renovation_Advance_Amount_Used__c -> Advance__c.Renovation_Reserve_Total__c -> Property__c.Approved_Renovation_Holdback__c -> Opp.Total_Amount_Advances__c
            adv_reno_val = None
            for a in advances:
                _f, adv_reno_val = pick_first_nonblank_field(a, ["Renovation_Reserve_Total__c"])
                if adv_reno_val is not None:
                    break
            total_reno_val = pick_first(
                prop.get("Renovation_Advance_Amount_Used__c"),
                adv_reno_val,
                prop.get("Approved_Renovation_Holdback__c"),
                opp.get("Total_Amount_Advances__c"),
            )
            sf_total_reno = parse_money(total_reno_val)

            # Interest Reserve: Property__c.Interest_Allocation__c -> Opp Interest_Reserves__c -> Opp Current_Interest_Reserves_Remaining__c -> Advance__c Interest_Reserve_Total__c -> Advance__c Total_Interest_Reserves_andStub_Interest__c
            adv_int_val = None
            for a in advances:
                _f, adv_int_val = pick_first_nonblank_field(
                    a,
                    ["Interest_Reserve_Total__c", "Total_Interest_Reserves_andStub_Interest__c", "Interest_Reserve_Subtotal__c"],
                )
                if adv_int_val is not None:
                    break
            interest_reserve_val = pick_first(
                prop.get("Interest_Allocation__c"),
                opp.get("Interest_Reserves__c"),
                opp.get("Current_Interest_Reserves_Remaining__c"),
                opp.get("Current_Interest_Reserves_Paid__c"),
                adv_int_val,
            )
            sf_interest_reserve = parse_money(interest_reserve_val)

            # Borrower + Address (prefer Property__c)
            borrower_final = (st.session_state.get("inp_borrower_disp") or "").strip().upper()
            address_final = (st.session_state.get("inp_address_disp") or "").strip().upper()

            # Deal # (Loan ID cell)
            deal_number_final = normalize_text(opp.get("Deal_Loan_Number__c")) or normalize_text(payload.get("deal_number")) or normalize_text(deal_input)

            # -----------------------------
            # Build ctx
            # -----------------------------
            ctx = {
                "deal_number": deal_number_final,
                "total_loan_amount": float(sf_total_loan_amount),
                "initial_advance": float(sf_initial_advance),
                "total_reno_drawn": float(sf_total_reno),
                "interest_reserve": float(sf_interest_reserve),

                "advance_amount": float(advance_amount),
                "holdback_pct": hb,
                "advance_date": st.session_state["inp_advance_date"].strftime("%m/%d/%Y"),

                "borrower_disp": borrower_final,
                "address_disp": address_final,

                "inspection_fee": float(inspection_fee),
                "wire_fee": float(wire_fee),
                "construction_mgmt_fee": float(construction_mgmt_fee),
                "title_fee": float(title_fee),
            }

            # Preview with source transparency (helps you validate fallbacks quickly)
            st.markdown("### Preview")
            prev = pd.DataFrame(
                [
                    ["Deal # (Loan ID cell)", ctx["deal_number"]],
                    ["Total Loan Amount", fmt_money(ctx["total_loan_amount"])],
                    ["Initial Advance", fmt_money(ctx["initial_advance"])],
                    ["Total Reno Drawn", fmt_money(ctx["total_reno_drawn"])],
                    ["Interest Reserve", fmt_money(ctx["interest_reserve"])],
                    ["Advance Amount", fmt_money(ctx["advance_amount"])],
                    ["Advance Date", ctx["advance_date"]],
                ],
                columns=["Field", "Value"],
            )
            st.dataframe(prev, use_container_width=True, hide_index=True)

            try:
                xbytes = build_hud_excel_bytes_from_template(ctx)
            except Exception as e:
                st.error("Could not build the HUD from the template.")
                st.code(str(e))
                st.stop()

            out_name = f"HUD_{re.sub(r'[^0-9A-Za-z_-]+','_', ctx['deal_number'] or 'Deal')}.xlsx"
            st.download_button(
                "Download HUD Excel",
                data=xbytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# -----------------------------
# CONSTRUCTION CHECKLIST
# -----------------------------
CHECKLIST_NOT_FOUND = "Not found"
CHECKLIST_TEMPLATE_CANDIDATES = [
    "Copy of Draw Check List REV 12.30.25 - Jonathan.xlsx",
    "Draw Check List REV 12.30.25.xlsx",
    "Draw Check List.xlsx",
]
CHECKLIST_STATUS_OPTIONS = ["Pending", "Complete", "Review", "Missing", "Not Applicable"]
CHECKLIST_EXPORT_SPECS = [
    {"order": 1, "field": "sold_loan_status", "label": "NLB (No Loan Balance) or Cap Partner / Sold Loan", "row_number": 2},
    {"order": 2, "field": "next_payment_due", "label": "Next payment due date", "row_number": 3},
    {"order": 3, "field": "late_payment_check", "label": "Other payments late?", "row_number": 4},
    {"order": 4, "field": "maturity_date", "label": "Maturity date", "row_number": 5},
    {"order": 5, "field": "tax_status", "label": "Taxes not delinquent", "row_number": 6},
    {"order": 6, "field": "supplier_code", "label": "Workday vendor set up / Supplier code", "row_number": 7},
    {"order": 7, "field": "insurance_status", "label": "Property insurance current", "row_number": 8},
]

FCI_DEFAULT_URL = "https://fapi.myfci.com/graphql"
FCI_LOAN_INFORMATION_QUERY = """
query GetLoanInformation {
  getLoanInformation {
    loanAccount: lenderAccount
    poffUnpaidLateCharges
    lateChargesDays
    lateChargesPct
    maturityDate
    nextDueDate
    noteRate
  }
}
"""


def get_fci_config() -> dict:
    try:
        cfg = st.secrets.get("fci", {})
    except Exception:
        cfg = {}
    url = normalize_text(cfg.get("url")) if hasattr(cfg, "get") else ""
    token = normalize_text(cfg.get("api_token")) if hasattr(cfg, "get") else ""
    return {
        "enabled": bool(token),
        "url": url or FCI_DEFAULT_URL,
        "api_token": token,
    }


@st.cache_data(ttl=300, show_spinner=False)
def fetch_fci_loan_information_rows(url: str, api_token: str) -> dict:
    if not url or not api_token:
        return {"ok": False, "rows": [], "error": "FCI API token is not configured."}
    headers = {"Authorization": f"Bearer {api_token}", "Content-Type": "application/json"}
    payload = {"query": FCI_LOAN_INFORMATION_QUERY, "variables": {}}
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
    except Exception as exc:
        return {"ok": False, "rows": [], "error": f"FCI loan information request failed: {exc}"}
    if "errors" in result:
        return {"ok": False, "rows": [], "error": json.dumps(result["errors"], indent=2)}
    records = (result.get("data") or {}).get("getLoanInformation") or []
    if isinstance(records, dict):
        records = [records]
    if not isinstance(records, list):
        records = []
    rows = []
    for record in records:
        if not isinstance(record, dict):
            continue
        rows.append(
            {
                "loanAccount": record.get("loanAccount"),
                "lenderAccount": record.get("loanAccount"),
                "poffUnpaidLateCharges": record.get("poffUnpaidLateCharges"),
                "lateChargesDays": record.get("lateChargesDays"),
                "lateChargesPct": record.get("lateChargesPct"),
                "maturityDate": record.get("maturityDate"),
                "nextDueDate": record.get("nextDueDate"),
                "noteRate": record.get("noteRate"),
            }
        )
    return {"ok": True, "rows": rows, "error": ""}


FCI_BORROWER_PAYMENT_QUERY = """
query GetBorrowerPayment {
  getBorrowerPayment {
    loanAccount: account
    dateReceived
    dateDue
    dayVariance
    paymentType
    totalAmount
    accruedLateCharges
    lateChargesPaid
    notes
    uid
  }
}
"""


def checklist_display_or_not_found(value) -> str:
    text = normalize_text(value)
    return text if text else CHECKLIST_NOT_FOUND


def checklist_yes_no(flag) -> str:
    if flag is True:
        return "Yes"
    if flag is False:
        return "No"
    return CHECKLIST_NOT_FOUND


def checklist_is_red_font(cell) -> bool:
    color = getattr(getattr(cell, "font", None), "color", None)
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return False
    return "FF0000" in str(rgb).upper()


def pick_checklist_template_bytes(uploaded_file) -> Tuple[bytes | None, str | None]:
    if uploaded_file is not None:
        return uploaded_file.getvalue(), uploaded_file.name
    for candidate in CHECKLIST_TEMPLATE_CANDIDATES:
        for base in [APP_DIR, Path('/mnt/data')]:
            path = base / candidate
            if path.exists():
                return path.read_bytes(), path.name
    return None, None


@st.cache_data(show_spinner=False)
def extract_checklist_template_rows(template_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]
    section = "General"
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        a = ws[f"A{row_idx}"]
        b = ws[f"B{row_idx}"]
        label = "" if a.value is None else str(a.value).strip()
        helper = "" if b.value is None else str(b.value).strip()
        if not label and not helper:
            continue
        is_section = bool(label and a.font and a.font.bold)
        if is_section:
            section = label
            continue
        rows.append(
            {
                "row_number": row_idx,
                "section": section,
                "item": label,
                "helper": helper,
                "is_red": checklist_is_red_font(a) or checklist_is_red_font(b),
                "status": "Pending",
                "value": "",
            }
        )
    return pd.DataFrame(rows)


def derive_expected_next_payment_due_from_close_date(close_date_value):
    close_dt = parse_date_any(close_date_value)
    if not close_dt:
        return ""
    cutoff = date(2025, 7, 1)
    due_day = 1 if close_dt > cutoff else 10
    today = date.today()
    year = today.year
    month = today.month
    if today.day > due_day:
        month += 1
        if month == 13:
            month = 1
            year += 1
    try:
        return date(year, month, due_day).strftime("%m/%d/%Y")
    except Exception:
        return ""


def fetch_account_by_id(account_id: str):
    if not account_id:
        return None
    fields = ["Id", "Name", "Phone", "Website"]
    where = f"Id = {soql_quote(account_id)}"
    rows, _used, _soql = try_query_drop_missing(sf, "Account", fields, where, limit=1)
    if not rows:
        return None
    row = rows[0].copy()
    row.pop("attributes", None)
    return row


def fetch_business_entity_by_id(entity_id: str):
    if not entity_id:
        return None
    fields = ["Id", "Name", "Borrower_Email_Address__c", "Operating_Agreement_Date__c"]
    where = f"Id = {soql_quote(entity_id)}"
    rows, _used, _soql = try_query_drop_missing(sf, "Business_Entity__c", fields, where, limit=1)
    if not rows:
        return None
    row = rows[0].copy()
    row.pop("attributes", None)
    return row


def fetch_checklist_opportunity_by_deal_number(deal_number: str):
    dn_digits = digits_only((deal_number or "").strip())
    if not dn_digits:
        return None
    fields = [
        "Id", "Name", "Deal_Loan_Number__c", "AccountId", "Borrower_Entity__c",
        "Intended_Capital_Partner__c", "Updated_Loan_Maturity_Date__c", "Next_Payment_Date__c",
        "CloseDate", "Servicer_Commitment_Id__c", "Warehouse_Line__c",
    ]
    where = (
        "("
        f"Deal_Loan_Number__c = {soql_quote(dn_digits)}"
        f" OR Deal_Loan_Number__c LIKE {soql_quote('%' + dn_digits + '%')}"
        ")"
    )
    rows, _used, _soql = try_query_drop_missing(sf, "Opportunity", fields, where, limit=10, order_by="CloseDate DESC")
    if not rows:
        return None
    row = rows[0].copy()
    row.pop("attributes", None)
    return row


def fetch_checklist_properties_for_deal(opp_id: str):
    lk = choose_first_existing("Property__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId"])
    if not lk:
        return []
    fields = [
        "Id", "Name", lk, "Property_Name__c", "Full_Address__c", "Next_Payment_Date__c",
        "Updated_Asset_Maturity_Date__c", "Servicer_Id__c", "ConstructionManagementLoanId__c",
        "Warehouse_Line_New__c", "Warehouse_Line__c",
    ]
    where = f"{lk} = {soql_quote(opp_id)}"
    rows, _used, _soql = try_query_drop_missing(sf, "Property__c", fields, where, limit=25, order_by="CreatedDate DESC")
    cleaned = []
    for row in rows:
        rec = row.copy()
        rec.pop("attributes", None)
        cleaned.append(rec)
    return cleaned


def fetch_servicer_loans_for_deal(opp_id: str):
    lk = choose_first_existing("Servicer_Loan__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId"])
    if not lk:
        return []
    fields = [
        "Id", "Name", lk, "Servicer_Commitment_ID__c", "Servicer_Loan_Status__c",
        "Delinquent_30_Days__c", "Delinquent_60_Days__c", "Delinquent_90_Days__c", "Delinquent_120_Days__c",
        "First_Payment_Date__c", "Last_Payment_Date__c",
    ]
    where = f"{lk} = {soql_quote(opp_id)}"
    rows, _used, _soql = try_query_drop_missing(sf, "Servicer_Loan__c", fields, where, limit=25, order_by="CreatedDate DESC")
    cleaned = []
    for row in rows:
        rec = row.copy()
        rec.pop("attributes", None)
        cleaned.append(rec)
    return cleaned


def fetch_sold_loan_pools_for_deal(opp_id: str):
    lk = choose_first_existing("Sold_Loan_Pool__c", ["Deal__c", "Opportunity__c", "Deal_Id__c", "OpportunityId", "DealId"])
    if not lk:
        return []
    fields = ["Id", "Name", lk, "Sold_To__c", "Status__c", "Servicing_Status__c", "Sold_Date__c"]
    where = f"{lk} = {soql_quote(opp_id)}"
    rows, _used, _soql = try_query_drop_missing(sf, "Sold_Loan_Pool__c", fields, where, limit=25, order_by="CreatedDate DESC")
    cleaned = []
    for row in rows:
        rec = row.copy()
        rec.pop("attributes", None)
        cleaned.append(rec)
    return cleaned


def _parse_float(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def _delinquency_found(servicer_row: dict):
    if not servicer_row:
        return None
    vals = [
        _parse_float(servicer_row.get("Delinquent_30_Days__c")),
        _parse_float(servicer_row.get("Delinquent_60_Days__c")),
        _parse_float(servicer_row.get("Delinquent_90_Days__c")),
        _parse_float(servicer_row.get("Delinquent_120_Days__c")),
    ]
    seen = any(v is not None for v in vals)
    if seen:
        return any((v or 0) > 0 for v in vals)
    status_text = normalize_text(servicer_row.get("Servicer_Loan_Status__c")).lower()
    if status_text:
        if any(token in status_text for token in ["delinq", "late", "default", "past due"]):
            return True
        return False
    return None


@st.cache_data(ttl=300, show_spinner=False)
def fetch_fci_borrower_payment_rows(url: str, api_token: str) -> dict:
    if not url or not api_token:
        return {"ok": False, "rows": [], "error": "FCI API token is not configured."}
    headers = {"Authorization": f"Bearer {api_token}", "Content-Type": "application/json"}
    payload = {"query": FCI_BORROWER_PAYMENT_QUERY, "variables": {}}
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
    except Exception as exc:
        return {"ok": False, "rows": [], "error": f"FCI borrower payment request failed: {exc}"}
    if "errors" in result:
        return {"ok": False, "rows": [], "error": json.dumps(result["errors"], indent=2)}
    records = (result.get("data") or {}).get("getBorrowerPayment") or []
    if isinstance(records, dict):
        records = [records]
    if not isinstance(records, list):
        records = []
    rows = []
    for record in records:
        if not isinstance(record, dict):
            continue
        rows.append(
            {
                "loanAccount": record.get("loanAccount"),
                "dateReceived": record.get("dateReceived"),
                "dateDue": record.get("dateDue"),
                "dayVariance": record.get("dayVariance"),
                "paymentType": record.get("paymentType"),
                "totalAmount": record.get("totalAmount"),
                "accruedLateCharges": record.get("accruedLateCharges"),
                "lateChargesPaid": record.get("lateChargesPaid"),
                "notes": record.get("notes"),
                "uid": record.get("uid"),
            }
        )
    return {"ok": True, "rows": rows, "error": ""}


def _fci_key(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", normalize_text(value)).upper()


def _group_rows_by_keys(rows: list[dict], field_names: list[str]) -> dict[str, list[dict]]:
    grouped = {}
    for row in rows or []:
        raw = ""
        for field_name in field_names:
            raw = normalize_text(row.get(field_name))
            if raw:
                break
        clean = _fci_key(raw)
        if not clean:
            continue
        grouped.setdefault(clean, []).append(row)
    return grouped


def _find_matching_fci_key(candidate_clean: str, available_keys: set[str]) -> str:
    if not candidate_clean:
        return ""
    if candidate_clean in available_keys:
        return candidate_clean
    cand_digits = digits_only(candidate_clean)
    for key in available_keys:
        if key.endswith(candidate_clean) or candidate_clean.endswith(key):
            return key
        key_digits = digits_only(key)
        if cand_digits and key_digits and cand_digits == key_digits:
            return key
    return ""


def get_checklist_servicer_key(bundle: dict) -> str:
    opp = bundle.get("opportunity") or {}
    prop = bundle.get("primary_property") or {}
    servicer_loans = bundle.get("servicer_loans") or []
    servicer_commitment = ""
    for row in servicer_loans:
        servicer_commitment = normalize_text(row.get("Servicer_Commitment_ID__c"))
        if servicer_commitment:
            break
    return pick_first(
        prop.get("Servicer_Id__c"),
        opp.get("Servicer_Commitment_Id__c"),
        prop.get("ConstructionManagementLoanId__c"),
        servicer_commitment,
    )


def build_fci_candidate_keys(bundle: dict) -> list[dict]:
    opp = bundle.get("opportunity") or {}
    prop = bundle.get("primary_property") or {}
    servicer_loans = bundle.get("servicer_loans") or []
    candidates = []
    def add(raw_value, label: str):
        raw_value = normalize_text(raw_value)
        clean_value = _fci_key(raw_value)
        if not clean_value:
            return
        if any(item["clean"] == clean_value for item in candidates):
            return
        candidates.append({"raw": raw_value, "clean": clean_value, "label": label})
    add(prop.get("Servicer_Id__c"), "Property servicer ID")
    add(opp.get("Servicer_Commitment_Id__c"), "Opportunity servicer commitment ID")
    add(prop.get("ConstructionManagementLoanId__c"), "Construction management loan ID")
    for idx, row in enumerate(servicer_loans, start=1):
        add(row.get("Servicer_Commitment_ID__c"), f"Servicer loan ID {idx}")
        add(row.get("Name"), f"Servicer loan name {idx}")
    return candidates


def fetch_fci_bundle(bundle: dict, loan_account_override: str = "") -> dict:
    cfg = get_fci_config()
    out = {
        "enabled": cfg.get("enabled", False),
        "candidate_keys": build_fci_candidate_keys(bundle),
        "matched": False,
        "match_source": "",
        "loan_account": "",
        "loan_info_record": {},
        "payment_rows": [],
        "loan_info_rows_found": 0,
        "payment_rows_found": 0,
        "error": "",
    }
    if not out["enabled"]:
        out["error"] = "FCI is not configured."
        return out
    loan_info_result = fetch_fci_loan_information_rows(cfg["url"], cfg["api_token"])
    payment_result = fetch_fci_borrower_payment_rows(cfg["url"], cfg["api_token"])
    loan_info_rows = loan_info_result.get("rows") or []
    payment_rows = payment_result.get("rows") or []
    out["loan_info_rows_found"] = len(loan_info_rows)
    out["payment_rows_found"] = len(payment_rows)
    loan_info_by_key = _group_rows_by_keys(loan_info_rows, ["loanAccount", "lenderAccount"])
    payment_by_key = _group_rows_by_keys(payment_rows, ["loanAccount", "account"])
    available_keys = set(loan_info_by_key.keys()) | set(payment_by_key.keys())

    def assign_match(matched_key: str, label: str) -> bool:
        if not matched_key:
            return False
        out["matched"] = True
        out["loan_account"] = matched_key
        out["match_source"] = label
        out["loan_info_record"] = (loan_info_by_key.get(matched_key) or [{}])[0]
        out["payment_rows"] = payment_by_key.get(matched_key) or []
        return True

    override_key = _find_matching_fci_key(_fci_key(loan_account_override), available_keys)
    if override_key and assign_match(override_key, "Manual FCI loan account override"):
        return out
    for candidate in out["candidate_keys"]:
        matched_key = _find_matching_fci_key(candidate["clean"], available_keys)
        if matched_key and assign_match(matched_key, candidate["label"]):
            return out
    if len(available_keys) == 1:
        only_key = next(iter(available_keys))
        if assign_match(only_key, "Single FCI loan account returned"):
            return out
    errors = []
    if not loan_info_result.get("ok"):
        errors.append(loan_info_result.get("error") or "Loan information query failed.")
    if not payment_result.get("ok"):
        errors.append(payment_result.get("error") or "Borrower payment query failed.")
    if not errors:
        errors.append("Could not match the Salesforce servicer ID to an FCI loan account.")
    out["error"] = " | ".join(errors)
    return out


def _late_payment_from_fci_loan_info(record: dict):
    if not record:
        return None
    amount = _parse_float(record.get("poffUnpaidLateCharges"))
    days = _parse_float(record.get("lateChargesDays"))
    pct = _parse_float(record.get("lateChargesPct"))
    if not any(v is not None for v in [amount, days, pct]):
        return None
    return any((v or 0) > 0 for v in [amount, days, pct])


def _late_payment_from_borrower_payments(rows: list[dict]):
    if not rows:
        return None
    usable = []
    for row in rows:
        payment_type = normalize_text(row.get("paymentType")).lower()
        if payment_type and payment_type not in {"regpmt", "regularpayment", "regular payment"}:
            continue
        notes = normalize_text(row.get("notes")).lower()
        if "reversed by" in notes:
            continue
        usable.append(row)
    if not usable:
        return None
    usable.sort(
        key=lambda r: (
            parse_date_any(r.get("dateReceived")) or date.min,
            parse_date_any(r.get("dateDue")) or date.min,
        ),
        reverse=True,
    )
    latest = usable[0]
    metrics = [
        _parse_float(latest.get("dayVariance")),
        _parse_float(latest.get("accruedLateCharges")),
        _parse_float(latest.get("lateChargesPaid")),
    ]
    if not any(v is not None for v in metrics):
        return None
    return any((v or 0) > 0 for v in metrics)


def _lookup_checklist_insurance(bundle: dict) -> dict:
    servicer_key = get_checklist_servicer_key(bundle)
    osc = osc_lookup(servicer_key)
    if not osc.get("found"):
        return {"servicer_key": servicer_key, "found": False, "value": CHECKLIST_NOT_FOUND, "error": normalize_text(osc.get("error")), "row": {}}
    row = osc.get("row") or {}
    primary_status = normalize_text(row.get("primary_status")).strip().lower()
    if not primary_status:
        value = CHECKLIST_NOT_FOUND
    else:
        value = "Yes" if primary_status == TARGET_INSURANCE_OK else "No"
    return {"servicer_key": servicer_key, "found": True, "value": value, "error": "", "row": row}


def _infer_tax_status_from_caf_row(row: dict) -> str:
    if not row:
        return CHECKLIST_NOT_FOUND
    preferred_cols = []
    for col in row.keys():
        lower = (col or "").lower()
        if any(token in lower for token in ["tax", "delinq", "delinquent", "install", "payment_status", "status"]):
            preferred_cols.append(col)
    for col in preferred_cols:
        raw = normalize_text(row.get(col))
        low = raw.lower()
        if not low:
            continue
        if any(token in low for token in ["not delinquent", "current", "paid", "clear", "good standing"]):
            return "Yes"
        if any(token in low for token in ["delinq", "delinquent", "late", "past due", "unpaid", "default"]):
            return "No"
        if "delinq" in (col or "").lower():
            if low in {"n", "no", "false", "0"}:
                return "Yes"
            if low in {"y", "yes", "true", "1"}:
                return "No"
    statuses = pick_payment_statuses(row)
    if statuses:
        return "Yes" if all(is_payment_status_ok(v) for (_k, v) in statuses) else "No"
    return CHECKLIST_NOT_FOUND


def _lookup_checklist_tax(bundle: dict, osc_match: dict) -> dict:
    opp = bundle.get("opportunity") or {}
    prop = bundle.get("primary_property") or {}
    deal_digits = digits_only(normalize_text(opp.get("Deal_Loan_Number__c")))
    caf = caf_try_match_by_deal_id(deal_digits)
    if not caf.get("found"):
        osc_addr = ""
        if osc_match.get("found"):
            row = osc_match.get("row") or {}
            osc_addr = " ".join([
                normalize_text(row.get("property_street")),
                normalize_text(row.get("property_city")),
                normalize_text(row.get("property_state")),
                normalize_text(row.get("property_zip")),
            ]).strip()
        caf = caf_try_match_by_address(normalize_text(prop.get("Full_Address__c")), osc_addr)
    tax_value = _infer_tax_status_from_caf_row(caf.get("row")) if caf.get("found") else CHECKLIST_NOT_FOUND
    return {"found": bool(caf.get("found")), "method": normalize_text(caf.get("method")), "value": tax_value, "error": normalize_text(caf.get("error")), "row": caf.get("row") or {}}


def derive_checklist_export_values(bundle: dict) -> Dict[str, str]:
    values = {spec["field"]: CHECKLIST_NOT_FOUND for spec in CHECKLIST_EXPORT_SPECS}
    if not bundle:
        return values
    opp = bundle.get("opportunity") or {}
    prop = bundle.get("primary_property") or {}
    sold_loan_pools = bundle.get("sold_loan_pools") or []
    servicer_loans = bundle.get("servicer_loans") or []
    fci = bundle.get("fci") or {}
    loan_info_record = fci.get("loan_info_record") or {}

    sold_flag = bool(sold_loan_pools or opp.get("Intended_Capital_Partner__c"))
    values["sold_loan_status"] = checklist_yes_no(sold_flag)

    next_due_value = pick_first(
        fmt_date_mmddyyyy(loan_info_record.get("nextDueDate")),
        fmt_date_mmddyyyy(prop.get("Next_Payment_Date__c")),
        fmt_date_mmddyyyy(opp.get("Next_Payment_Date__c")),
        derive_expected_next_payment_due_from_close_date(opp.get("CloseDate")),
    )
    values["next_payment_due"] = checklist_display_or_not_found(next_due_value)

    late_flag = _late_payment_from_borrower_payments(fci.get("payment_rows") or [])
    if late_flag is None:
        late_flag = _late_payment_from_fci_loan_info(loan_info_record)
    if late_flag is None:
        for servicer_row in servicer_loans:
            late_flag = _delinquency_found(servicer_row)
            if late_flag is not None:
                break
    values["late_payment_check"] = checklist_yes_no(late_flag)

    maturity_value = pick_first(
        fmt_date_mmddyyyy(loan_info_record.get("maturityDate")),
        fmt_date_mmddyyyy(prop.get("Updated_Asset_Maturity_Date__c")),
        fmt_date_mmddyyyy(opp.get("Updated_Loan_Maturity_Date__c")),
    )
    values["maturity_date"] = checklist_display_or_not_found(maturity_value)

    insurance_lookup = _lookup_checklist_insurance(bundle)
    values["insurance_status"] = insurance_lookup.get("value") or CHECKLIST_NOT_FOUND

    tax_lookup = _lookup_checklist_tax(bundle, insurance_lookup)
    values["tax_status"] = tax_lookup.get("value") or CHECKLIST_NOT_FOUND

    supplier_code = pick_first(
        prop.get("Warehouse_Line_New__c"),
        prop.get("Warehouse_Line__c"),
        opp.get("Warehouse_Line__c"),
    )
    values["supplier_code"] = checklist_display_or_not_found(supplier_code)

    bundle["_checklist_support"] = {
        "servicer_key": insurance_lookup.get("servicer_key") or get_checklist_servicer_key(bundle),
        "fci_match_source": normalize_text(fci.get("match_source")),
        "fci_loan_account": normalize_text(fci.get("loan_account")),
        "fci_payment_rows": len(fci.get("payment_rows") or []),
        "fci_error": normalize_text(fci.get("error")),
        "tax_found": bool(tax_lookup.get("found")),
        "tax_method": normalize_text(tax_lookup.get("method")),
        "tax_error": normalize_text(tax_lookup.get("error")),
        "insurance_found": bool(insurance_lookup.get("found")),
        "insurance_error": normalize_text(insurance_lookup.get("error")),
    }
    return values


def build_checklist_export_rows(export_values: dict) -> pd.DataFrame:
    rows = []
    for spec in CHECKLIST_EXPORT_SPECS:
        rows.append(
            {
                "order": spec["order"],
                "field": spec["field"],
                "checklist_item": spec["label"],
                "value": checklist_display_or_not_found(export_values.get(spec["field"])),
            }
        )
    return pd.DataFrame(rows)


def build_checklist_export_excel_bytes(export_df: pd.DataFrame, deal_number: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist Values"
    ws.append(["Deal Number", "Checklist Item", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FF000000")
    for _, row in export_df.iterrows():
        ws.append([deal_number, row["checklist_item"], row["value"]])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 28
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_checklist_export_summary(export_df: pd.DataFrame):
    found = int((export_df["value"] != CHECKLIST_NOT_FOUND).sum())
    missing = int((export_df["value"] == CHECKLIST_NOT_FOUND).sum())
    c1, c2 = st.columns(2)
    c1.metric("Values found", found)
    c2.metric("Not found", missing)


def build_checklist_auto_answers(form_values: dict) -> Dict[int, dict]:
    answers = {}
    def clean_value(field_name: str) -> str:
        value = checklist_display_or_not_found(form_values.get(field_name))
        return "" if value == CHECKLIST_NOT_FOUND else value
    sold_value = checklist_display_or_not_found(form_values.get("sold_loan_status"))
    answers[2] = {"status": "Complete" if sold_value != CHECKLIST_NOT_FOUND else "Review", "value": "" if sold_value == CHECKLIST_NOT_FOUND else sold_value}
    due_value = checklist_display_or_not_found(form_values.get("next_payment_due"))
    answers[3] = {"status": "Complete" if due_value != CHECKLIST_NOT_FOUND else "Review", "value": clean_value("next_payment_due")}
    late_value = checklist_display_or_not_found(form_values.get("late_payment_check"))
    answers[4] = {"status": "Complete" if late_value == "No" else "Review", "value": "" if late_value == CHECKLIST_NOT_FOUND else late_value}
    maturity_value = checklist_display_or_not_found(form_values.get("maturity_date"))
    answers[5] = {"status": "Complete" if maturity_value != CHECKLIST_NOT_FOUND else "Review", "value": clean_value("maturity_date")}
    tax_value = checklist_display_or_not_found(form_values.get("tax_status"))
    answers[6] = {"status": "Complete" if tax_value == "Yes" else "Review", "value": "" if tax_value == CHECKLIST_NOT_FOUND else tax_value}
    supplier_value = checklist_display_or_not_found(form_values.get("supplier_code"))
    answers[7] = {"status": "Complete" if supplier_value != CHECKLIST_NOT_FOUND else "Review", "value": clean_value("supplier_code")}
    insurance_value = checklist_display_or_not_found(form_values.get("insurance_status"))
    answers[8] = {"status": "Complete" if insurance_value == "Yes" else "Review", "value": "" if insurance_value == CHECKLIST_NOT_FOUND else insurance_value}
    return answers


def apply_checklist_auto_answers(base_df: pd.DataFrame, answers: Dict[int, dict]) -> pd.DataFrame:
    df = base_df.copy()
    for row_number, payload in answers.items():
        mask = df["row_number"] == row_number
        if not mask.any():
            continue
        df.loc[mask, "status"] = payload.get("status", "")
        df.loc[mask, "value"] = payload.get("value", "")
    return df


def build_checklist_output_workbook(template_bytes: bytes, edited_rows: pd.DataFrame) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]
    ws["C1"] = "Status"
    ws["D1"] = "Value"
    header_font = Font(bold=True, color="FF000000")
    ws["C1"].font = header_font
    ws["D1"].font = header_font
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 36
    for _, row in edited_rows.iterrows():
        r = int(row["row_number"])
        ws[f"C{r}"] = row["status"]
        ws[f"D{r}"] = row["value"]
        ws[f"C{r}"].font = Font(color="FF000000")
        ws[f"D{r}"].font = Font(color="FF000000")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def fetch_construction_checklist_bundle(deal_number: str, loan_account_override: str = ""):
    opp = fetch_checklist_opportunity_by_deal_number(deal_number)
    if not opp:
        return None
    opp_id = opp.get("Id")
    account = fetch_account_by_id(opp.get("AccountId")) if opp.get("AccountId") else None
    business_entity = fetch_business_entity_by_id(opp.get("Borrower_Entity__c")) if opp.get("Borrower_Entity__c") else None
    cap_partner_account = fetch_account_by_id(opp.get("Intended_Capital_Partner__c")) if opp.get("Intended_Capital_Partner__c") else None
    properties = fetch_checklist_properties_for_deal(opp_id)
    primary_property = properties[0] if properties else None
    servicer_loans = fetch_servicer_loans_for_deal(opp_id)
    sold_loan_pools = fetch_sold_loan_pools_for_deal(opp_id)
    sold_to_account = None
    if sold_loan_pools and sold_loan_pools[0].get("Sold_To__c"):
        sold_to_account = fetch_account_by_id(sold_loan_pools[0].get("Sold_To__c"))
    bundle = {
        "opportunity": opp,
        "account": account,
        "business_entity": business_entity,
        "cap_partner_account": cap_partner_account,
        "properties": properties,
        "primary_property": primary_property,
        "servicer_loans": servicer_loans,
        "sold_loan_pools": sold_loan_pools,
        "sold_to_account": sold_to_account,
    }
    bundle["fci"] = fetch_fci_bundle(bundle, loan_account_override)
    return bundle


def run_construction_checklist_page():
    ensure_default("checklist_deal_number_input", "")
    ensure_default("checklist_loan_account_override", "")
    ensure_default("checklist_bundle", None)
    ensure_default("checklist_export_values", {})

    st.subheader("Construction Checklist")
    st.caption("Enter a deal number to pull the checklist values. The screen only shows the information the user needs.")

    with st.expander("Admin options", expanded=False):
        uploaded_template = st.file_uploader(
            "Upload the construction checklist workbook",
            type=["xlsx"],
            key="construction_template_upload",
        )
        st.text_input(
            "FCI loan account override",
            key="checklist_loan_account_override",
            placeholder="Use only if the automatic FCI match misses",
        )
        st.write("OSC file:", osc_path_used, "✅" if osc_err is None else "❌")
        st.write("CAF tax file:", caf_path_used, "✅" if caf_err is None else "❌")

    template_bytes, template_name = pick_checklist_template_bytes(uploaded_template)

    c1, c2 = st.columns([2.5, 1.0])
    with c1:
        deal_input = st.text_input(
            "Deal Number",
            key="checklist_deal_number_input",
            placeholder="Enter the deal number",
        )
    with c2:
        pull_btn = st.button("Get checklist values", type="primary", use_container_width=True)

    if pull_btn:
        st.session_state.checklist_bundle = None
        st.session_state.checklist_export_values = {}
        bundle = fetch_construction_checklist_bundle(
            deal_input,
            st.session_state.get("checklist_loan_account_override", ""),
        )
        if not bundle:
            st.error("No deal found for that number.")
        else:
            values = derive_checklist_export_values(bundle)
            st.session_state.checklist_bundle = bundle
            st.session_state.checklist_export_values = values

    bundle = st.session_state.get("checklist_bundle")
    export_values = st.session_state.get("checklist_export_values") or {}
    if not bundle:
        st.info("Enter a deal number and click Get checklist values.")
        return

    opp = bundle.get("opportunity") or {}
    prop = bundle.get("primary_property") or {}
    account = bundle.get("account") or {}
    st.markdown(
        f"""
<div class="soft-card">
  <div class="big"><b>{normalize_text(opp.get('Deal_Loan_Number__c')) or normalize_text(st.session_state.get('checklist_deal_number_input'))}</b></div>
  <div class="muted">{normalize_text(opp.get('Name')) or 'Deal'} • {normalize_text(account.get('Name')) or 'Borrower not found'}</div>
  <div class="muted">{normalize_text(prop.get('Property_Name__c') or prop.get('Name')) or 'Property not found'}</div>
  <div class="muted">{normalize_text(prop.get('Full_Address__c')) or 'Address not found'}</div>
</div>
""",
        unsafe_allow_html=True,
    )

    export_df = build_checklist_export_rows(export_values)
    st.markdown("### Checklist values")
    render_checklist_export_summary(export_df)
    editor_df = export_df[["checklist_item", "value"]].copy()
    edited_df = st.data_editor(
        editor_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=["checklist_item"],
        key="construction_checklist_values_editor",
        column_config={
            "checklist_item": st.column_config.TextColumn("Checklist item", width="large", disabled=True),
            "value": st.column_config.TextColumn("Value", width="medium"),
        },
    )
    edited_export_df = export_df.copy()
    edited_export_df["value"] = edited_df["value"].astype(str)

    edited_values = {}
    for spec in CHECKLIST_EXPORT_SPECS:
        mask = edited_export_df["field"] == spec["field"]
        value = CHECKLIST_NOT_FOUND
        if mask.any():
            value = checklist_display_or_not_found(edited_export_df.loc[mask, "value"].iloc[0])
        edited_values[spec["field"]] = value

    deal_number_for_file = normalize_text(opp.get("Deal_Loan_Number__c")) or normalize_text(st.session_state.get("checklist_deal_number_input")) or "deal"
    export_csv = edited_export_df[["checklist_item", "value"]].to_csv(index=False).encode("utf-8")
    export_xlsx = build_checklist_export_excel_bytes(edited_export_df[["checklist_item", "value"]], deal_number_for_file)
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "Download checklist values (CSV)",
            data=export_csv,
            file_name=f"construction_checklist_values_{re.sub(r'[^0-9A-Za-z_-]+', '_', deal_number_for_file)}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "Download checklist values (Excel)",
            data=export_xlsx,
            file_name=f"construction_checklist_values_{re.sub(r'[^0-9A-Za-z_-]+', '_', deal_number_for_file)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if template_bytes is not None:
        base_df = extract_checklist_template_rows(template_bytes)
        working_df = apply_checklist_auto_answers(base_df, build_checklist_auto_answers(edited_values))
        output_bytes = build_checklist_output_workbook(template_bytes, working_df)
        st.download_button(
            "Download completed checklist workbook",
            data=output_bytes,
            file_name=f"construction_checklist_completed_{re.sub(r'[^0-9A-Za-z_-]+', '_', deal_number_for_file)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        st.caption(f"Workbook template: {template_name}")
    else:
        st.info("Upload the checklist workbook in Admin options if you want the completed workbook export.")

    support = bundle.get("_checklist_support") or {}
    with st.expander("Troubleshooting details", expanded=False):
        st.write("FCI match source:", support.get("fci_match_source") or "Not found")
        st.write("FCI loan account:", support.get("fci_loan_account") or "Not found")
        st.write("FCI payment rows:", support.get("fci_payment_rows", 0))
        if support.get("fci_error"):
            st.write("FCI note:", support.get("fci_error"))
        st.write("Servicer ID used:", support.get("servicer_key") or "Not found")
        st.write("Tax match method:", support.get("tax_method") or "Not found")
        if support.get("tax_error"):
            st.write("Tax note:", support.get("tax_error"))
        if support.get("insurance_error"):
            st.write("Insurance note:", support.get("insurance_error"))


def run_app():
    workflow = st.sidebar.radio("Workflow", ["HUD Generator", "Construction Checklist"], index=0)
    if workflow == "HUD Generator":
        run_hud_generator_page()
    else:
        run_construction_checklist_page()


if __name__ == "__main__":
    run_app()
